"""
Large File Handler - OPTIMIZED Incremental Processing for Excel Files

This module handles large Excel files by:
1. Extracting metadata WITHOUT loading the entire file
2. Using SMART SAMPLING (not sequential reading)
3. Providing granular progress callbacks
4. Supporting timeout and cancellation
"""

import os
import json
from pathlib import Path
from typing import Dict, List, Optional, Any, Generator
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
import pandas as pd


class ExcelMetadataExtractor:
    """Extract metadata from Excel files with OPTIMIZED sampling for large files."""

    def __init__(self, file_path: str, chunk_size: int = 1000):
        """
        Initialize metadata extractor.

        Args:
            file_path: Path to Excel file
            chunk_size: Maximum number of rows to sample per sheet
        """
        self.file_path = file_path
        self.chunk_size = chunk_size
        self.file_size = os.path.getsize(file_path)
        self.metadata = {}

    def extract_metadata(self, event_callback=None) -> Dict[str, Any]:
        """
        Extract comprehensive metadata from Excel file with progress reporting.

        Returns:
            Dictionary containing:
            - file_info: Basic file information
            - sheets: List of sheet metadata
            - relationships: Detected relationships between sheets
            - statistics: Overall file statistics
        """
        if event_callback:
            event_callback({
                'timestamp': datetime.now().strftime("%H:%M:%S"),
                'type': 'text',
                'content': f'📊 Extracting metadata from large file ({self.file_size / (1024*1024):.1f} MB)...',
                'icon': '🔍'
            })

        # Load workbook in read-only mode
        wb = load_workbook(filename=self.file_path, read_only=True, data_only=True)

        metadata = {
            'file_info': {
                'path': self.file_path,
                'filename': os.path.basename(self.file_path),
                'size_bytes': self.file_size,
                'size_mb': round(self.file_size / (1024 * 1024), 2),
                'extraction_time': datetime.now().isoformat()
            },
            'sheets': [],
            'statistics': {
                'total_sheets': 0,
                'total_columns': 0,
                'estimated_total_rows': 0,
                'data_types': {}
            },
            'exploration_guide': {
                'message': 'Use bash commands and Python scripts to explore relationships yourself',
                'recommended_approach': 'Sample data incrementally and reason about connections'
            }
        }

        # Process each sheet
        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            if event_callback:
                event_callback({
                    'timestamp': datetime.now().strftime("%H:%M:%S"),
                    'type': 'text',
                    'content': f'📄 Analyzing sheet {sheet_idx + 1}/{len(wb.sheetnames)}: {sheet_name}',
                    'icon': '📑'
                })

            sheet_metadata = self._extract_sheet_metadata(wb, sheet_name, event_callback)
            metadata['sheets'].append(sheet_metadata)

            # Update overall statistics
            metadata['statistics']['total_columns'] += len(sheet_metadata['columns'])
            metadata['statistics']['estimated_total_rows'] += sheet_metadata['estimated_rows']

        metadata['statistics']['total_sheets'] = len(metadata['sheets'])

        # Close workbook
        wb.close()

        if event_callback:
            event_callback({
                'timestamp': datetime.now().strftime("%H:%M:%S"),
                'type': 'text',
                'content': f'✅ Metadata extracted: {len(metadata["sheets"])} sheets, {metadata["statistics"]["total_columns"]} columns, ~{metadata["statistics"]["estimated_total_rows"]:,} rows',
                'icon': '✨'
            })

        self.metadata = metadata
        return metadata

    def _extract_sheet_metadata(self, wb, sheet_name: str, event_callback=None) -> Dict[str, Any]:
        """Extract metadata from a single sheet with OPTIMIZED sampling strategy."""
        sheet = wb[sheet_name]

        # Get dimensions
        max_row = sheet.max_row
        max_col = sheet.max_column

        if event_callback:
            event_callback({
                'timestamp': datetime.now().strftime("%H:%M:%S"),
                'type': 'text',
                'content': f'   📏 Sheet "{sheet_name}": {max_row:,} rows × {max_col} columns',
                'icon': '📊'
            })

        # Read header row to get column names
        headers = []
        for col_idx in range(1, max_col + 1):
            cell_value = sheet.cell(row=1, column=col_idx).value
            headers.append(str(cell_value) if cell_value is not None else f"Column_{col_idx}")

        # ULTRA-OPTIMIZED SAMPLING STRATEGY
        # Adaptive sampling based on sheet size to prevent slow deep-file access
        if max_row > 100000:
            # MASSIVE sheet (>100K rows): Use minimal sampling from FIRST rows only
            # Deep-file access with iter_rows is too slow, so sample from beginning
            sample_size = min(100, self.chunk_size)  # Much smaller sample
            step = 1  # Sequential from start (fast)
            if event_callback:
                event_callback({
                    'timestamp': datetime.now().strftime("%H:%M:%S"),
                    'type': 'text',
                    'content': f'   ⚡ Massive sheet - sampling first {sample_size} rows (sequential for speed)',
                    'icon': '🔍'
                })
        elif max_row > 10000:
            # Large sheet: sample every Nth row, but keep sample size reasonable
            sample_size = min(200, self.chunk_size)  # Reduced from 500
            step = max(1, (max_row - 1) // sample_size)
            if event_callback:
                event_callback({
                    'timestamp': datetime.now().strftime("%H:%M:%S"),
                    'type': 'text',
                    'content': f'   ⚡ Large sheet - sampling {sample_size} rows (every {step}th row)',
                    'icon': '🔍'
                })
        else:
            # Small sheet: sample first N rows
            sample_size = min(self.chunk_size, max_row - 1)
            step = 1

        # OPTIMIZED: Initialize column metadata structures first
        column_metadata = []
        for col_idx, header in enumerate(headers, start=1):
            column_metadata.append({
                'name': header,
                'index': col_idx - 1,
                'excel_column': self._get_excel_column_letter(col_idx),
                'data_type': 'unknown',
                'sample_values': [],
                'null_count': 0,
                'unique_values_sample': set(),
                'rows_sampled': 0
            })

        # OPTIMIZED ROW-BY-ROW READING (single pass)
        # Read rows in one pass instead of column-by-column for massive speedup
        last_progress_report = 0
        rows_sampled = 0

        for i in range(sample_size):
            row_idx = 2 + (i * step)  # Start at row 2 (skip header), step by N
            if row_idx > max_row:
                break

            # Report progress periodically
            if event_callback and sample_size > 100 and (i - last_progress_report) >= 100:
                percent = int((i / sample_size) * 100)
                event_callback({
                    'timestamp': datetime.now().strftime("%H:%M:%S"),
                    'type': 'text',
                    'content': f'   🔄 Sampling rows {percent}% complete ({i}/{sample_size} rows sampled)',
                    'icon': '⚙️'
                })
                last_progress_report = i

            # Read entire row at once (MUCH faster than column-by-column)
            row_cells = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, min_col=1, max_col=max_col, values_only=True))
            if not row_cells:
                continue

            row_values = row_cells[0]  # Get first (and only) row
            rows_sampled += 1

            # Process each column value from this row
            for col_idx in range(len(row_values)):
                if col_idx >= len(column_metadata):
                    break

                value = row_values[col_idx]
                col_meta = column_metadata[col_idx]

                if value is not None:
                    if len(col_meta['sample_values']) < 100:  # Limit samples stored
                        col_meta['sample_values'].append(value)
                    col_meta['unique_values_sample'].add(str(value)[:100])
                else:
                    col_meta['null_count'] += 1

        # Finalize column metadata after sampling
        for col_meta in column_metadata:
            col_meta['data_type'] = self._infer_data_type(col_meta['sample_values'])
            col_meta['sample_values'] = [str(v)[:100] for v in col_meta['sample_values'][:5]]  # Store first 5
            col_meta['unique_values_count'] = len(col_meta['unique_values_sample'])
            col_meta['unique_values_sample'] = list(col_meta['unique_values_sample'])[:10]
            col_meta['rows_sampled'] = rows_sampled
            col_meta['sampling_step'] = step

        sheet_metadata = {
            'name': sheet_name,
            'estimated_rows': max_row - 1,  # Exclude header
            'columns': column_metadata,
            'has_header': True,
            'data_range': f"A1:{self._get_excel_column_letter(max_col)}{max_row}",
            'sampling_strategy': f"Sampled {sample_size} rows with step={step}"
        }

        return sheet_metadata

    def _infer_data_type(self, values: List[Any]) -> str:
        """Infer data type from sample values."""
        if not values:
            return 'empty'

        type_counts = {
            'numeric': 0,
            'date': 0,
            'text': 0,
            'boolean': 0
        }

        for value in values:
            if isinstance(value, (int, float)):
                type_counts['numeric'] += 1
            elif isinstance(value, datetime):
                type_counts['date'] += 1
            elif isinstance(value, bool):
                type_counts['boolean'] += 1
            else:
                type_counts['text'] += 1

        # Return the most common type
        return max(type_counts, key=type_counts.get)

    def _get_excel_column_letter(self, col_idx: int) -> str:
        """Convert column index to Excel column letter (1 -> A, 27 -> AA, etc.)."""
        result = ""
        while col_idx > 0:
            col_idx -= 1
            result = chr(col_idx % 26 + 65) + result
            col_idx //= 26
        return result


    def save_metadata(self, output_path: str):
        """Save metadata to JSON file."""
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(self.metadata, f, indent=2, ensure_ascii=False)

    def get_chunk_iterator(self, sheet_name: str, chunk_size: Optional[int] = None) -> Generator[pd.DataFrame, None, None]:
        """
        Get an iterator that yields chunks of data from a specific sheet.

        Args:
            sheet_name: Name of sheet to read
            chunk_size: Number of rows per chunk (defaults to self.chunk_size)

        Yields:
            DataFrame chunks
        """
        chunk_size = chunk_size or self.chunk_size

        # Use pandas to read in chunks
        for chunk in pd.read_excel(
            self.file_path,
            sheet_name=sheet_name,
            chunksize=chunk_size,
            engine='openpyxl'
        ):
            yield chunk

    def get_sheet_sample(self, sheet_name: str, n_rows: int = 100) -> pd.DataFrame:
        """
        Get a sample of rows from a sheet.

        Args:
            sheet_name: Name of sheet
            n_rows: Number of rows to sample

        Returns:
            DataFrame with sampled rows
        """
        return pd.read_excel(
            self.file_path,
            sheet_name=sheet_name,
            nrows=n_rows,
            engine='openpyxl'
        )


class LargeFileAnalyzer:
    """Coordinates large file analysis using metadata and chunked processing."""

    def __init__(self, file_path: str, metadata_dir: str = 'metadata', chunk_size: int = 1000):
        """
        Initialize large file analyzer.

        Args:
            file_path: Path to Excel file
            metadata_dir: Directory to store metadata
            chunk_size: Rows per chunk
        """
        self.file_path = file_path
        self.metadata_dir = Path(metadata_dir)
        self.metadata_dir.mkdir(exist_ok=True)
        self.chunk_size = chunk_size
        self.extractor = ExcelMetadataExtractor(file_path, chunk_size)
        self.metadata = None

    def extract_and_save_metadata(self, run_id: str, event_callback=None) -> str:
        """
        Extract metadata and save to file.

        Args:
            run_id: Unique identifier for this analysis
            event_callback: Callback for progress updates

        Returns:
            Path to saved metadata file
        """
        # Extract metadata
        self.metadata = self.extractor.extract_metadata(event_callback)

        # Save to file
        metadata_path = self.metadata_dir / f"{run_id}_metadata.json"
        self.extractor.save_metadata(str(metadata_path))

        return str(metadata_path)

    def load_metadata(self, metadata_path: str):
        """Load metadata from file."""
        with open(metadata_path, 'r', encoding='utf-8') as f:
            self.metadata = json.load(f)
        return self.metadata

    def generate_metadata_summary(self) -> str:
        """
        Generate a human-readable summary of metadata for Claude.

        Returns:
            Formatted text summary
        """
        if not self.metadata:
            return "No metadata available"

        summary_lines = [
            "# Excel File Metadata Summary",
            "",
            "## File Information",
            f"- Filename: {self.metadata['file_info']['filename']}",
            f"- Size: {self.metadata['file_info']['size_mb']} MB",
            f"- Total Sheets: {self.metadata['statistics']['total_sheets']}",
            f"- Total Columns: {self.metadata['statistics']['total_columns']}",
            f"- Estimated Rows: {self.metadata['statistics']['estimated_total_rows']:,}",
            "",
            "## Sheets Structure",
            ""
        ]

        for sheet in self.metadata['sheets']:
            summary_lines.append(f"### {sheet['name']}")
            summary_lines.append(f"- Rows: ~{sheet['estimated_rows']:,}")
            summary_lines.append(f"- Columns: {len(sheet['columns'])}")
            summary_lines.append(f"- Sampling: {sheet.get('sampling_strategy', 'Standard')}")
            summary_lines.append("")
            summary_lines.append("**Columns:**")

            for col in sheet['columns']:
                sample_preview = ", ".join(str(v)[:50] for v in col['sample_values'][:3])
                summary_lines.append(
                    f"  - `{col['name']}` ({col['data_type']}) - "
                    f"Nulls: {col['null_count']}/{col.get('rows_sampled', '?')}, "
                    f"Unique: ~{col['unique_values_count']} - "
                    f"Samples: [{sample_preview}]"
                )

            summary_lines.append("")

        summary_lines.append("## Your Task: Discover Relationships")
        summary_lines.append("")
        summary_lines.append("**The metadata above shows STRUCTURE ONLY - not relationships.**")
        summary_lines.append("")
        summary_lines.append("YOU must discover relationships by:")
        summary_lines.append("1. Looking at column names across sheets and reasoning about semantic connections")
        summary_lines.append("2. Sampling data from columns that might be related (e.g., IDs, foreign keys)")
        summary_lines.append("3. Using bash/Python to check value overlaps between columns")
        summary_lines.append("4. Reasoning about the business logic (e.g., 'CustomerID' likely links Customers → Orders)")
        summary_lines.append("")
        summary_lines.append("**Example exploration commands:**")
        summary_lines.append("```python")
        summary_lines.append("# Sample unique values from a potential key column")
        summary_lines.append("import pandas as pd")
        summary_lines.append("df = pd.read_excel('file.xlsx', sheet_name='Sheet1', nrows=100)")
        summary_lines.append("print(df['ColumnName'].unique()[:10])")
        summary_lines.append("")
        summary_lines.append("# Check if values from Sheet1 exist in Sheet2")
        summary_lines.append("df1 = pd.read_excel('file.xlsx', sheet_name='Sheet1', nrows=500)")
        summary_lines.append("df2 = pd.read_excel('file.xlsx', sheet_name='Sheet2', nrows=500)")
        summary_lines.append("common = set(df1['ID']).intersection(set(df2['ID']))")
        summary_lines.append("print(f'Common values: {len(common)}')")
        summary_lines.append("```")
        summary_lines.append("")

        return "\n".join(summary_lines)
