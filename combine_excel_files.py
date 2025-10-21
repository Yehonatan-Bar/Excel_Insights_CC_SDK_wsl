#!/usr/bin/env python3
"""
Script to combine multiple Excel files into a single workbook with multiple sheets.
Each source file becomes a separate sheet in the output file.
"""

import os
from pathlib import Path
from copy import copy
import openpyxl
from openpyxl.utils import get_column_letter

def combine_excel_files(source_dir, output_file):
    """
    Combine all Excel files from source_dir into a single workbook.

    Args:
        source_dir: Directory containing the Excel files to combine
        output_file: Path to the output Excel file
    """
    # Create a new workbook for the combined output
    output_wb = openpyxl.Workbook()
    # Remove the default sheet created
    output_wb.remove(output_wb.active)

    # Get all Excel files from the source directory
    source_path = Path(source_dir)
    excel_files = sorted(source_path.glob('*.xlsx'))

    if not excel_files:
        print(f"No Excel files found in {source_dir}")
        return

    print(f"Found {len(excel_files)} Excel file(s) to combine:")

    for excel_file in excel_files:
        print(f"  - Processing: {excel_file.name}")

        try:
            # Load the source workbook
            source_wb = openpyxl.load_workbook(excel_file, data_only=False)

            # Process each sheet in the source workbook
            for sheet_name in source_wb.sheetnames:
                source_sheet = source_wb[sheet_name]

                # Create a unique sheet name in the output workbook
                # Use filename (without extension) + original sheet name
                base_name = excel_file.stem
                if len(source_wb.sheetnames) == 1:
                    # If there's only one sheet, just use the filename
                    new_sheet_name = base_name[:31]  # Excel sheet name limit is 31 chars
                else:
                    # If multiple sheets, combine filename and sheet name
                    new_sheet_name = f"{base_name}_{sheet_name}"[:31]

                # Make sure the sheet name is unique
                counter = 1
                original_name = new_sheet_name
                while new_sheet_name in output_wb.sheetnames:
                    new_sheet_name = f"{original_name[:28]}_{counter}"
                    counter += 1

                # Create the new sheet in output workbook
                new_sheet = output_wb.create_sheet(title=new_sheet_name)

                # Copy all cells from source sheet to new sheet
                for row in source_sheet.iter_rows():
                    for cell in row:
                        new_cell = new_sheet[cell.coordinate]

                        # Copy cell value
                        new_cell.value = cell.value

                        # Copy cell formatting
                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.border = copy(cell.border)
                            new_cell.fill = copy(cell.fill)
                            new_cell.number_format = cell.number_format
                            new_cell.protection = copy(cell.protection)
                            new_cell.alignment = copy(cell.alignment)

                # Copy column dimensions
                for col in source_sheet.column_dimensions:
                    if col in source_sheet.column_dimensions:
                        new_sheet.column_dimensions[col].width = \
                            source_sheet.column_dimensions[col].width

                # Copy row dimensions
                for row in source_sheet.row_dimensions:
                    if row in source_sheet.row_dimensions:
                        new_sheet.row_dimensions[row].height = \
                            source_sheet.row_dimensions[row].height

                # Copy merged cells
                for merged_cell_range in source_sheet.merged_cells.ranges:
                    new_sheet.merge_cells(str(merged_cell_range))

                print(f"    > Added sheet: '{new_sheet_name}'")

            source_wb.close()

        except Exception as e:
            print(f"    X Error processing {excel_file.name}: {str(e)}")
            continue

    # Save the combined workbook
    print(f"\nSaving combined workbook to: {output_file}")
    output_wb.save(output_file)
    output_wb.close()

    print(f"+ Successfully combined {len(excel_files)} files into {len(output_wb.sheetnames)} sheet(s)")
    print(f"\nOutput file created: {output_file}")
    print(f"Total sheets in output: {len(output_wb.sheetnames)}")

if __name__ == "__main__":
    # Define paths
    script_dir = Path(__file__).parent
    source_directory = script_dir / "sources"
    output_filename = script_dir / "Combined_Excel_Files.xlsx"

    print("=" * 60)
    print("Excel Files Combiner")
    print("=" * 60)
    print(f"Source directory: {source_directory}")
    print(f"Output file: {output_filename}")
    print("=" * 60)
    print()

    # Combine the files
    combine_excel_files(source_directory, output_filename)

    print()
    print("=" * 60)
    print("Process completed!")
    print("=" * 60)
